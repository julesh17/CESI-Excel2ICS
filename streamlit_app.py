import streamlit as st
import pandas as pd
import re
import uuid
from datetime import datetime, date, time
from dateutil import parser as dtparser
import pytz
from openpyxl import load_workbook
import io

# ---------------- Utilities ----------------

def normalize_group_label(x):
    if x is None:
        return None
    try:
        if pd.isna(x):
            return None
    except Exception:
        pass
    s = str(x).strip()
    if not s:
        return None
    m = re.search(r'G\s*\.?\s*(\d+)', s, re.I)
    if m:
        return f'G {m.group(1)}'
    m2 = re.search(r'^(?:groupe)?\s*(\d+)$', s, re.I)
    if m2:
        return f'G {m2.group(1)}'
    return s


def is_time_like(x):
    if x is None:
        return False
    if isinstance(x, (pd.Timestamp, datetime, time)):
        return True
    s = str(x).strip()
    if not s:
        return False
    # accepte formats hh:mm, hhm m, 9h30, 9:30, 9AM, 9 PM ...
    if re.match(r'^\d{1,2}[:hH]\d{2}(\s*[AaPp][Mm]\.?)*$', s):
        return True
    return False


def to_time(x):
    if x is None:
        return None
    if isinstance(x, time):
        return x
    if isinstance(x, pd.Timestamp):
        return x.to_pydatetime().time()
    if isinstance(x, datetime):
        return x.time()
    s = str(x).strip()
    if not s:
        return None
    s2 = s.replace('h', ':').replace('H', ':')
    try:
        dt = dtparser.parse(s2, dayfirst=True)
        return dt.time()
    except Exception:
        return None


def to_date(x):
    if x is None:
        return None
    if isinstance(x, pd.Timestamp):
        return x.to_pydatetime().date()
    if isinstance(x, datetime):
        return x.date()
    if isinstance(x, date):
        return x
    s = str(x).strip()
    if not s:
        return None
    try:
        dt = dtparser.parse(s, dayfirst=True, fuzzy=True)
        return dt.date()
    except Exception:
        return None

# ---------------- Helpers Fusion ----------------

def get_merged_map(xls_fileobj, sheet_name):
    """
    xls_fileobj: file-like (BytesIO or path)
    Retourne un dict {(row0,col0): (r1,c1,r2,c2)} où les indices sont 0-based
    """
    wb = load_workbook(xls_fileobj, data_only=True)
    ws = wb[sheet_name]
    merged_map = {}
    for merged in ws.merged_cells.ranges:
        r1, r2 = merged.min_row, merged.max_row
        c1, c2 = merged.min_col, merged.max_col
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                merged_map[(r - 1, c - 1)] = (r1 - 1, c1 - 1, r2 - 1, c2 - 1)
    return merged_map

# ---------------- Parsing ----------------

def find_week_rows(df):
    return [i for i in range(len(df)) if isinstance(df.iat[i, 0], str) and re.match(r'^\s*S\s*\d+', df.iat[i, 0].strip(), re.I)]


def find_slot_rows(df):
    return [i for i in range(len(df)) if isinstance(df.iat[i, 0], str) and re.match(r'^\s*H\s*\d+', df.iat[i, 0].strip(), re.I)]


def parse_sheet_to_events(xls_fileobj, sheet_name):
    """
    xls_fileobj: file-like (BytesIO). sheet_name: string.
    Retour: liste d'événements dict {summary, teachers, description, start, end, groups}
    """
    # pandas accepte un BytesIO
    df = pd.read_excel(xls_fileobj, sheet_name=sheet_name, header=None)
    nrows, ncols = df.shape

    # pour load_workbook on doit récréer un BytesIO car pandas a consommé le flux
    # si xls_fileobj est BytesIO, getbuffer() / seek(0) possible
    # Ici on reopen via a new BytesIO from the original content if available
    if hasattr(xls_fileobj, "getvalue"):
        book_io_for_openpyxl = io.BytesIO(xls_fileobj.getvalue())
    else:
        # fallback (si xls_fileobj est un path)
        book_io_for_openpyxl = xls_fileobj

    merged_map = get_merged_map(book_io_for_openpyxl, sheet_name)

    s_rows = find_week_rows(df)
    h_rows = find_slot_rows(df)

    raw_events = []

    for r in h_rows:
        p_candidates = [s for s in s_rows if s <= r]
        if not p_candidates:
            continue
        p = max(p_candidates)
        date_row = p + 1
        group_row = p + 2

        date_cols = [c for c in range(ncols) if date_row < nrows and to_date(df.iat[date_row, c]) is not None]

        for c in date_cols:
            for col in (c, c + 1):
                if col >= ncols:
                    continue
                summary = df.iat[r, col] if col < ncols else None
                if pd.isna(summary) or summary is None:
                    continue
                summary_str = str(summary).strip()
                if not summary_str:
                    continue

                # --- Teachers ---
                teachers = []
                if (r + 2) < nrows:
                    for off in range(2, 6):
                        idx = r + off
                        if idx >= nrows:
                            break
                        t = df.iat[idx, col]
                        if t is None or pd.isna(t):
                            continue
                        s = str(t).strip()
                        if not s:
                            continue
                        if not is_time_like(s) and to_date(s) is None:
                            teachers.append(s)
                teachers = list(dict.fromkeys(teachers))

                # --- Stop index ---
                stop_idx = None
                for off in range(1, 12):
                    idx = r + off
                    if idx >= nrows:
                        break
                    if is_time_like(df.iat[idx, col]):
                        stop_idx = idx
                        break
                if stop_idx is None:
                    stop_idx = min(r + 7, nrows)

                # --- Description ---
                desc_parts = []
                for idx in range(r + 1, stop_idx):
                    if idx >= nrows:
                        break
                    cell = df.iat[idx, col]
                    if pd.isna(cell) or cell is None:
                        continue
                    s = str(cell).strip()
                    if not s:
                        continue
                    if to_date(cell) is not None:
                        continue
                    if s in teachers or s == summary_str:
                        continue
                    desc_parts.append(s)
                desc_text = " | ".join(dict.fromkeys(desc_parts))

                # --- Heures ---
                start_val, end_val = None, None
                for off in range(1, 13):
                    idx = r + off
                    if idx >= nrows:
                        break
                    v = df.iat[idx, col]
                    if is_time_like(v):
                        if start_val is None:
                            start_val = v
                        elif end_val is None and v != start_val:
                            end_val = v
                            break
                if start_val is None or end_val is None:
                    continue
                start_t, end_t = to_time(start_val), to_time(end_val)
                if start_t is None or end_t is None:
                    continue

                # --- Date ---
                d = to_date(df.iat[date_row, c])
                if d is None:
                    continue
                dtstart, dtend = datetime.combine(d, start_t), datetime.combine(d, end_t)

                # --- Groupes ---
                gl = normalize_group_label(df.iat[group_row, col] if group_row < nrows else None)
                gl_next = normalize_group_label(df.iat[group_row, col + 1] if (col + 1) < ncols else None)
                is_left_col = (col == c)
                right_summary = df.iat[r, col + 1] if (col + 1) < ncols else None

                groups = set()
                if is_left_col:
                    # Vérif fusion réelle dans Excel
                    merged = merged_map.get((r, col))
                    # si fusion détectée couvrant la cellule et la suivante, ajouter les 2 groupes
                    if merged and (r, col + 1) in merged_map:
                        if gl:
                            groups.add(gl)
                        if gl_next:
                            groups.add(gl_next)
                    else:
                        if gl:
                            groups.add(gl)
                else:
                    if gl:
                        groups.add(gl)

                raw_events.append({
                    'summary': summary_str,
                    'teachers': set(teachers),
                    'descriptions': set([desc_text]) if desc_text else set(),
                    'start': dtstart,
                    'end': dtend,
                    'groups': groups
                })

    # --- Fusion des événements --- 
    merged = {}
    for e in raw_events:
        key = (e['summary'], e['start'], e['end'])
        if key not in merged:
            merged[key] = {
                'summary': e['summary'],
                'teachers': set(),
                'descriptions': set(),
                'start': e['start'],
                'end': e['end'],
                'groups': set()
            }
        merged[key]['teachers'].update(e.get('teachers', set()))
        merged[key]['descriptions'].update(e.get('descriptions', set()))
        merged[key]['groups'].update(e.get('groups', set()))

    return [ {
        'summary': v['summary'],
        'teachers': sorted(list(v['teachers'])),
        'description': " | ".join(sorted(list(v['descriptions']))) if v['descriptions'] else "",
        'start': v['start'],
        'end': v['end'],
        'groups': sorted(list(v['groups']))
    } for v in merged.values() ]

# ---------------- ICS writer ----------------

def escape_ical_text(s: str) -> str:
    if s is None:
        return ""
    s = str(s)
    s = s.replace('\\', '\\\\').replace('\n', '\\n').replace(',', '\\,').replace(';', '\\;')
    return s

def build_paris_vtimezone_text():
    """Retourne un bloc VTIMEZONE texte pour Europe/Paris (utile pour clients qui attendent VTIMEZONE)."""
    return "\n".join([
        "BEGIN:VTIMEZONE",
        "TZID:Europe/Paris",
        "X-LIC-LOCATION:Europe/Paris",
        "BEGIN:DAYLIGHT",
        "TZOFFSETFROM:+0100",
        "TZOFFSETTO:+0200",
        "TZNAME:CEST",
        "DTSTART:19700329T020000",
        "RRULE:FREQ=YEARLY;BYMONTH=3;BYDAY=-1SU",
        "END:DAYLIGHT",
        "BEGIN:STANDARD",
        "TZOFFSETFROM:+0200",
        "TZOFFSETTO:+0100",
        "TZNAME:CET",
        "DTSTART:19701025T030000",
        "RRULE:FREQ=YEARLY;BYMONTH=10;BYDAY=-1SU",
        "END:STANDARD",
        "END:VTIMEZONE"
    ])

def events_to_ics(events, tzname='Europe/Paris'):
    """
    events: list of dicts with keys 'summary','teachers','description','start','end','groups'
    Retourne un string ICS complet incluant VTIMEZONE.
    """
    tz = pytz.timezone(tzname)
    header = [
        'BEGIN:VCALENDAR',
        'VERSION:2.0',
        'PRODID:-//EDT Export//FR',
        'CALSCALE:GREGORIAN',
    ]

    # ajouter VTIMEZONE textuel (Outlook a besoin des règles pour correctement afficher DST)
    vtz = build_paris_vtimezone_text()
    body = [vtz]

    for ev in events:
        uid = str(uuid.uuid4())
        # localize pour avoir l'heure locale (naive -> tz aware)
        # if event times are already tz-aware, do not localize
        start_dt = ev['start']
        end_dt = ev['end']
        if start_dt.tzinfo is None:
            start_loc = tz.localize(start_dt)
        else:
            start_loc = start_dt.astimezone(tz)
        if end_dt.tzinfo is None:
            end_loc = tz.localize(end_dt)
        else:
            end_loc = end_dt.astimezone(tz)

        dtstart = start_loc.strftime('%Y%m%dT%H%M%S')
        dtend = end_loc.strftime('%Y%m%dT%H%M%S')
        summary = escape_ical_text(ev['summary'])

        desc_lines = []
        if ev.get('description'):
            desc_lines.append(ev['description'])
        if ev.get('teachers'):
            desc_lines.append('Enseignant(s): ' + ' / '.join(ev['teachers']))
        groups = ev.get('groups', [])
        if groups:
            if len(groups) == 1:
                desc_lines.append('Groupe: ' + groups[0])
            else:
                desc_lines.append('Groupes: ' + ' et '.join(groups))

        description = escape_ical_text('\n'.join(desc_lines))

        body.extend([
            'BEGIN:VEVENT',
            f'UID:{uid}',
            f'DTSTAMP:{datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")}',
            f'DTSTART;TZID={tzname}:{dtstart}',
            f'DTEND;TZID={tzname}:{dtend}',
            f'SUMMARY:{summary}',
            f'DESCRIPTION:{description}',
            'END:VEVENT'
        ])

    footer = ['END:VCALENDAR']
    return '\n'.join(header + body + footer)

# ---------------- Streamlit UI ----------------

st.set_page_config(page_title='Excel → ICS (EDT)', layout='centered')
st.title('Convertisseur Emplois du Temps (Excel → .ics)')

uploaded = st.file_uploader('Choisir le fichier Excel (.xlsx)', type=['xlsx'])
if uploaded is None:
    st.info('Upload un fichier .xlsx pour commencer.')
    st.stop()

# Lire les bytes une seule fois et créer un BytesIO réutilisable
file_bytes = uploaded.read()
file_io = io.BytesIO(file_bytes)

try:
    xls = pd.ExcelFile(io.BytesIO(file_bytes))
    sheets = xls.sheet_names
except Exception as e:
    st.error('Impossible de lire le fichier Excel: ' + str(e))
    st.stop()

st.write('Feuilles trouvées :', sheets)

for sheet in ['EDT P1', 'EDT P2']:
    if sheet in sheets:
        st.header(sheet)
        # on passe un BytesIO "frais" à la fonction (parse_sheet_to_events lira via pandas et openpyxl)
        events = parse_sheet_to_events(io.BytesIO(file_bytes), sheet)
        if not events:
            st.warning(f'Aucun événement détecté dans {sheet} (vérifier la structure).')
            continue

        ics = events_to_ics(events, tzname='Europe/Paris')
        st.write(f'{len(events)} événements extraits pour {sheet}')
        st.download_button(label=f'Télécharger {sheet}.ics', data=ics, file_name=f'{sheet}.ics', mime='text/calendar')
